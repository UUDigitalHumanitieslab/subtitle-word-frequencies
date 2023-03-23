# Parse the metadata xlsx file

from openpyxl import load_workbook

def parse(filename):
    wb = load_workbook(filename)
    sheet = wb[wb.sheetnames[0]]

    fieldnames = parse_header(sheet)
    datarows = parse_body(sheet, fieldnames)
    filtered = filter(is_not_empty, datarows) #ignore empty rows

    return filtered

def parse_header(sheet):
    header = next(sheet.rows)
    return cell_values(header)

def parse_body(sheet, fieldnames, min_row = 2): # start at the second row (index is from 1, not 0)
    return (
        parse_row(row, fieldnames)
        for row in
        sheet.iter_rows(min_row = min_row)
    )

def parse_row(row, fieldnames):
    return dict(zip(fieldnames, cell_values(row)))

def cell_values(cells):
    return [cell.value for cell in cells]

def is_not_empty(data):
    return any(value for value in data.values())
